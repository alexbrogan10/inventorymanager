import csv
import io
from collections.abc import Callable
from datetime import date, timedelta

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.user import UserRole

REPORTS_URL = "/api/v1/reports"
PRODUCTS_URL = "/api/v1/products"


def _auth_header(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _create_category(client: TestClient, token: str, name: str = "Electronics") -> int:
    response = client.post("/api/v1/categories", json={"name": name}, headers=_auth_header(token))
    return int(response.json()["id"])


def _create_supplier(client: TestClient, token: str, company_name: str = "Acme Supply Co.") -> int:
    response = client.post(
        "/api/v1/suppliers",
        json={
            "company_name": company_name,
            "contact_person": "Jane Doe",
            "email": "jane@acme.example",
            "phone": "555-0100",
            "address": "123 Warehouse Rd",
            "lead_time_days": 7,
            "notes": None,
        },
        headers=_auth_header(token),
    )
    return int(response.json()["id"])


def _create_warehouse(client: TestClient, token: str, name: str = "Main Warehouse") -> int:
    response = client.post(
        "/api/v1/warehouses",
        json={"name": name, "address": "Somewhere", "notes": None},
        headers=_auth_header(token),
    )
    return int(response.json()["id"])


def _create_product(
    client: TestClient,
    token: str,
    category_id: int,
    supplier_id: int,
    sku: str,
    purchase_price: str = "5.00",
    selling_price: str = "9.99",
    minimum_quantity: int = 10,
) -> int:
    response = client.post(
        PRODUCTS_URL,
        json={
            "sku": sku,
            "barcode": None,
            "name": f"Product {sku}",
            "description": None,
            "category_id": category_id,
            "supplier_id": supplier_id,
            "purchase_price": purchase_price,
            "selling_price": selling_price,
            "minimum_quantity": minimum_quantity,
            "maximum_quantity": None,
            "unit_type": "each",
        },
        headers=_auth_header(token),
    )
    return int(response.json()["id"])


def _set_level(
    client: TestClient, token: str, product_id: int, warehouse_id: int, quantity: int
) -> None:
    client.put(
        f"{PRODUCTS_URL}/{product_id}/inventory/{warehouse_id}",
        json={"quantity": quantity},
        headers=_auth_header(token),
    )


def _create_sale(
    client: TestClient,
    token: str,
    warehouse_id: int,
    product_id: int,
    quantity: int,
    unit_price: str,
) -> int:
    response = client.post(
        "/api/v1/sales",
        json={
            "warehouse_id": warehouse_id,
            "customer_name": "Report Customer",
            "items": [{"product_id": product_id, "quantity": quantity, "unit_price": unit_price}],
        },
        headers=_auth_header(token),
    )
    return int(response.json()["id"])


def _create_purchase_order(
    client: TestClient,
    token: str,
    supplier_id: int,
    warehouse_id: int,
    product_id: int,
    quantity_ordered: int = 5,
    unit_cost: str = "1.00",
    expected_delivery_date: str | None = None,
) -> dict:
    response = client.post(
        "/api/v1/purchase-orders",
        json={
            "supplier_id": supplier_id,
            "warehouse_id": warehouse_id,
            "expected_delivery_date": expected_delivery_date,
            "items": [
                {
                    "product_id": product_id,
                    "quantity_ordered": quantity_ordered,
                    "unit_cost": unit_cost,
                }
            ],
        },
        headers=_auth_header(token),
    )
    return dict(response.json())


def _ship_and_receive(client: TestClient, token: str, order_id: int) -> None:
    client.post(f"/api/v1/purchase-orders/{order_id}/ship", headers=_auth_header(token))
    client.post(f"/api/v1/purchase-orders/{order_id}/receive", headers=_auth_header(token))


def _transfer(
    client: TestClient,
    token: str,
    product_id: int,
    from_warehouse_id: int,
    to_warehouse_id: int,
    quantity: int,
) -> None:
    client.post(
        f"{PRODUCTS_URL}/{product_id}/transfer",
        json={
            "from_warehouse_id": from_warehouse_id,
            "to_warehouse_id": to_warehouse_id,
            "quantity": quantity,
        },
        headers=_auth_header(token),
    )


class Fixture:
    """A category/supplier/two-warehouse setup plus one product - enough for
    every report (transfers need a second warehouse to move stock into)."""

    def __init__(self, client: TestClient, auth_token_for: Callable[..., str]) -> None:
        self.token = auth_token_for(UserRole.MANAGER, email="reports-manager@example.com")
        category_id = _create_category(client, self.token)
        self.supplier_id = _create_supplier(client, self.token)
        self.warehouse_id = _create_warehouse(client, self.token, name="Main Warehouse")
        self.warehouse_b_id = _create_warehouse(client, self.token, name="Secondary Warehouse")
        self.product_id = _create_product(
            client, self.token, category_id, self.supplier_id, sku="PROD-A"
        )


class TestAccess:
    def test_requires_authentication(self, client: TestClient) -> None:
        assert client.get(f"{REPORTS_URL}/inventory-valuation").status_code == 401

    def test_any_authenticated_role_can_read(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        token = auth_token_for(UserRole.EMPLOYEE)

        response = client.get(f"{REPORTS_URL}/inventory-valuation", headers=_auth_header(token))

        assert response.status_code == 200


class TestInventoryValuation:
    def test_computes_value_at_cost_and_potential_revenue(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _set_level(client, fixture.token, fixture.product_id, fixture.warehouse_id, 20)

        response = client.get(
            f"{REPORTS_URL}/inventory-valuation", headers=_auth_header(fixture.token)
        )

        body = response.json()
        assert len(body["rows"]) == 1
        row = body["rows"][0]
        assert row["total_quantity"] == 20
        assert row["value_at_cost"] == "100.00"
        assert row["potential_revenue"] == "199.80"
        assert body["total_value_at_cost"] == "100.00"
        assert body["total_potential_revenue"] == "199.80"

    def test_product_with_no_inventory_rows_shows_zero_quantity(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)

        response = client.get(
            f"{REPORTS_URL}/inventory-valuation", headers=_auth_header(fixture.token)
        )

        row = response.json()["rows"][0]
        assert row["total_quantity"] == 0
        assert row["value_at_cost"] == "0.00"

    def test_csv_export_is_well_formed(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _set_level(client, fixture.token, fixture.product_id, fixture.warehouse_id, 20)

        response = client.get(
            f"{REPORTS_URL}/inventory-valuation",
            params={"format": "csv"},
            headers=_auth_header(fixture.token),
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")
        assert "inventory-valuation.csv" in response.headers["content-disposition"]
        rows = list(csv.DictReader(io.StringIO(response.text)))
        assert len(rows) == 1
        assert rows[0]["sku"] == "PROD-A"
        assert rows[0]["total_quantity"] == "20"

    def test_xlsx_export_is_well_formed(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _set_level(client, fixture.token, fixture.product_id, fixture.warehouse_id, 20)

        response = client.get(
            f"{REPORTS_URL}/inventory-valuation",
            params={"format": "xlsx"},
            headers=_auth_header(fixture.token),
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert "sku" in header
        data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert data_row[header.index("sku")] == "PROD-A"
        assert data_row[header.index("total_quantity")] == 20


class TestSalesHistory:
    def test_reports_item_count_and_revenue_ascending_by_date(
        self, client: TestClient, auth_token_for: Callable[..., str], db_session: Session
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _set_level(client, fixture.token, fixture.product_id, fixture.warehouse_id, 100)
        first_sale = _create_sale(
            client, fixture.token, fixture.warehouse_id, fixture.product_id, 2, "9.99"
        )
        second_sale = _create_sale(
            client, fixture.token, fixture.warehouse_id, fixture.product_id, 3, "9.99"
        )
        # Force apart timestamps that would otherwise tie within one test
        # transaction (Postgres's now() is constant per transaction) so
        # ascending order is actually exercised - see test_dashboard.py.
        db_session.execute(
            text("UPDATE sales SET created_at = created_at - interval '1 minute' WHERE id = :id"),
            {"id": first_sale},
        )
        db_session.commit()

        response = client.get(f"{REPORTS_URL}/sales-history", headers=_auth_header(fixture.token))

        body = response.json()
        assert [row["sale_id"] for row in body["rows"]] == [first_sale, second_sale]
        assert body["rows"][0]["item_count"] == 2
        assert body["rows"][0]["total_revenue"] == "19.98"
        assert body["rows"][0]["sold_by"] == "Test User"
        assert body["total_revenue"] == "49.95"

    def test_date_range_filters_out_of_range_sales(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _set_level(client, fixture.token, fixture.product_id, fixture.warehouse_id, 100)
        _create_sale(client, fixture.token, fixture.warehouse_id, fixture.product_id, 1, "9.99")
        tomorrow = (date.today() + timedelta(days=1)).isoformat()

        response = client.get(
            f"{REPORTS_URL}/sales-history",
            params={"start_date": tomorrow},
            headers=_auth_header(fixture.token),
        )

        assert response.json()["rows"] == []


class TestPurchaseHistory:
    def test_reports_item_count_cost_and_status(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)

        order = _create_purchase_order(
            client,
            fixture.token,
            fixture.supplier_id,
            fixture.warehouse_id,
            fixture.product_id,
            quantity_ordered=5,
            unit_cost="2.00",
        )

        response = client.get(
            f"{REPORTS_URL}/purchase-history", headers=_auth_header(fixture.token)
        )

        body = response.json()
        assert len(body["rows"]) == 1
        row = body["rows"][0]
        assert row["purchase_order_id"] == order["id"]
        assert row["item_count"] == 5
        assert row["total_cost"] == "10.00"
        assert row["status"] == "ordered"
        assert row["supplier"] == "Acme Supply Co."
        assert body["total_cost"] == "10.00"

    def test_date_range_filters_out_of_range_orders(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        _create_purchase_order(
            client, fixture.token, fixture.supplier_id, fixture.warehouse_id, fixture.product_id
        )
        yesterday = (date.today() - timedelta(days=1)).isoformat()

        response = client.get(
            f"{REPORTS_URL}/purchase-history",
            params={"end_date": yesterday},
            headers=_auth_header(fixture.token),
        )

        assert response.json()["rows"] == []


class TestProductMovement:
    def test_returns_404_for_unknown_product(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        token = auth_token_for(UserRole.EMPLOYEE)

        response = client.get(
            f"{REPORTS_URL}/product-movement",
            params={"product_id": 999999},
            headers=_auth_header(token),
        )

        assert response.status_code == 404

    def test_merges_receipts_sales_and_transfers_ascending(
        self, client: TestClient, auth_token_for: Callable[..., str], db_session: Session
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        order = _create_purchase_order(
            client,
            fixture.token,
            fixture.supplier_id,
            fixture.warehouse_id,
            fixture.product_id,
            quantity_ordered=50,
        )
        _ship_and_receive(client, fixture.token, order["id"])
        sale_id = _create_sale(
            client, fixture.token, fixture.warehouse_id, fixture.product_id, 4, "9.99"
        )
        _transfer(
            client,
            fixture.token,
            fixture.product_id,
            fixture.warehouse_id,
            fixture.warehouse_b_id,
            10,
        )

        # Force the receipt to be earliest and the sale to be latest so
        # ascending order is meaningfully exercised, mirroring the tie-break
        # workaround used elsewhere for same-transaction timestamps.
        db_session.execute(
            text(
                "UPDATE purchase_orders SET updated_at = updated_at - interval '2 minutes' "
                "WHERE id = :id"
            ),
            {"id": order["id"]},
        )
        db_session.execute(
            text("UPDATE sales SET created_at = created_at + interval '2 minutes' WHERE id = :id"),
            {"id": sale_id},
        )
        db_session.commit()

        response = client.get(
            f"{REPORTS_URL}/product-movement",
            params={"product_id": fixture.product_id},
            headers=_auth_header(fixture.token),
        )

        body = response.json()
        rows = body["rows"]
        # 1 receipt + 1 sale + 2 transfer legs (out + in) = 4 rows.
        assert len(rows) == 4
        types = [row["type"] for row in rows]
        assert types[0] == "purchase_receipt"
        assert types[-1] == "sale"
        assert {"transfer_out", "transfer_in"}.issubset(set(types))

        receipt_row = next(row for row in rows if row["type"] == "purchase_receipt")
        assert receipt_row["quantity_change"] == 50
        assert receipt_row["reference"] == f"PO #{order['id']}"

        sale_row = next(row for row in rows if row["type"] == "sale")
        assert sale_row["quantity_change"] == -4
        assert sale_row["reference"] == f"Sale #{sale_id}"

        transfer_out = next(row for row in rows if row["type"] == "transfer_out")
        transfer_in = next(row for row in rows if row["type"] == "transfer_in")
        assert transfer_out["quantity_change"] == -10
        assert transfer_out["warehouse"] == "Main Warehouse"
        assert transfer_in["quantity_change"] == 10
        assert transfer_in["warehouse"] == "Secondary Warehouse"

        for row in rows:
            assert row["product_id"] == fixture.product_id
            assert row["product_sku"] == "PROD-A"

    def test_only_returns_events_for_the_requested_product(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        other_product_id = _create_product(
            client,
            fixture.token,
            _create_category(client, fixture.token, name="Other"),
            fixture.supplier_id,
            sku="PROD-B",
        )
        _set_level(client, fixture.token, other_product_id, fixture.warehouse_id, 50)
        _create_sale(client, fixture.token, fixture.warehouse_id, other_product_id, 1, "9.99")

        response = client.get(
            f"{REPORTS_URL}/product-movement",
            params={"product_id": fixture.product_id},
            headers=_auth_header(fixture.token),
        )

        assert response.json()["rows"] == []


class TestSupplierPerformance:
    def test_computes_spend_and_on_time_rate_from_received_orders_only(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        fixture = Fixture(client, auth_token_for)
        on_time = _create_purchase_order(
            client,
            fixture.token,
            fixture.supplier_id,
            fixture.warehouse_id,
            fixture.product_id,
            quantity_ordered=10,
            unit_cost="2.00",
            expected_delivery_date=(date.today() + timedelta(days=5)).isoformat(),
        )
        _ship_and_receive(client, fixture.token, on_time["id"])
        late = _create_purchase_order(
            client,
            fixture.token,
            fixture.supplier_id,
            fixture.warehouse_id,
            fixture.product_id,
            quantity_ordered=5,
            unit_cost="3.00",
            expected_delivery_date=(date.today() - timedelta(days=5)).isoformat(),
        )
        _ship_and_receive(client, fixture.token, late["id"])
        cancelled = _create_purchase_order(
            client, fixture.token, fixture.supplier_id, fixture.warehouse_id, fixture.product_id
        )
        client.post(
            f"/api/v1/purchase-orders/{cancelled['id']}/cancel", headers=_auth_header(fixture.token)
        )

        response = client.get(
            f"{REPORTS_URL}/supplier-performance", headers=_auth_header(fixture.token)
        )

        rows = response.json()["rows"]
        assert len(rows) == 1
        row = rows[0]
        assert row["company_name"] == "Acme Supply Co."
        assert row["total_orders"] == 3
        assert row["total_received"] == 2
        assert row["total_cancelled"] == 1
        # (10 * 2.00) + (5 * 3.00) = 35.00, cancelled order excluded.
        assert row["total_spend"] == "35.00"
        assert row["average_lead_time_days"] == 0.0
        assert row["on_time_rate"] == 0.5

    def test_supplier_with_no_orders_is_omitted(
        self, client: TestClient, auth_token_for: Callable[..., str]
    ) -> None:
        token = auth_token_for(UserRole.MANAGER, email="reports-manager2@example.com")
        _create_supplier(client, token, company_name="Unused Supplier")

        response = client.get(f"{REPORTS_URL}/supplier-performance", headers=_auth_header(token))

        assert response.json()["rows"] == []
